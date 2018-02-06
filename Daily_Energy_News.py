import bs4
import requests
import openpyxl
import datetime
import ctypes

ctypes.windll.user32.MessageBoxW(0, "Start Web Scrape", "Daily Energy News", 1)

today = datetime.datetime.now()

wb = openpyxl.load_workbook(r'Daily Energy.xlsx')
sheet = wb.active


def forex():
    bloomURL = requests.get('http://www.bloomberg.com/markets/currencies/asia-pacific')

    soup = bs4.BeautifulSoup(bloomURL.text, "html.parser")

    for r, tr in enumerate(soup.find_all('tr')[11:12]):
        tds = tr.find_all('td')
        for c, x in enumerate(tds[1:2]):
            sheet.cell(row = 1, column = 2).value = float(x.text)
        
def index():
    bloomURL = requests.get('http://www.bloomberg.com/markets/stocks/world-indexes/asia-pacific')

    soup = bs4.BeautifulSoup(bloomURL.text)

    for r, tr in enumerate(soup.find_all('tr')[96:97]):
        tds = tr.find_all('td')
        for c, x in enumerate(tds[1:2]):
            sheet.cell(row = 2, column = 2).value = x.text

def stocks(url, row):
    bloomURL = requests.get(url)

    soup = bs4.BeautifulSoup(bloomURL.text)

    table = soup.find_all('div', {'class':'value__b93f12ea'})
    price_range_left = soup.find_all('span', {'class':'textLeft'})
    price_range_right = soup.find_all('span', {'class':'textRight'})
    price = soup.find_all('span', {'class':'priceText__1853e8a5'})
    change = soup.find_all('span', {'class':'changePercent__2d7dc0d2'})
    volume = soup.find_all('div', {'class':'value__b93f12ea'})

    try:
        sheet.cell(row = row, column = 2).value = float(table[1].text)
        sheet.cell(row = row, column = 3).value = float(table[0].text)
        sheet.cell(row = row, column = 4).value = price_range_left[0].text + '-' + price_range_right[0].text
        sheet.cell(row = row, column = 5).value = float(price[0].text)

        try:
            sheet.cell(row = row, column = 6).value = change[0].text
        except IndexError:
            sheet.cell(row = row, column = 6).value = '0.00%'

        sheet.cell(row = row, column = 7).value = table[2].text
        
    except (ValueError, IndexError):
        pass
        

def ngcp():
    ngcpURL = requests.get('http://ngcp.ph/')
    soup = bs4.BeautifulSoup(ngcpURL.text)
    table = soup.find_all('table', class_ = 'table table-bordered SmallText')
    td = table[0].find_all('td')

    #System Capacity
    luz_capacity = td[6]
    vis_capacity = td[7]
    min_capacity = td[8]

    #System Peak
    luz_peak = td[10]
    vis_peak = td[11]
    min_peak = td[12]

    #System Reserve
    luz_reserve = td[14]
    vis_reserve = td[15]
    min_reserve = td[16]

    try:
        sheet.cell(row = 20, column = 2).value = float(luz_capacity.string)
        sheet.cell(row = 21, column = 2).value = float(luz_peak.string)
        sheet.cell(row = 22, column = 2).value = float(luz_reserve.string)
    except ValueError:
        print('Luzon NGCP data missing')

    try:        
        sheet.cell(row = 20, column = 3).value = float(vis_capacity.string)
        sheet.cell(row = 21, column = 3).value = float(vis_peak.string)
        sheet.cell(row = 22, column = 3).value = float(vis_reserve.string)
    except ValueError:
        print('Visayas NGCP data missing')

    try:
        sheet.cell(row = 20, column = 4).value = float(min_capacity.string)
        sheet.cell(row = 21, column = 4).value = float(min_peak.string)
        sheet.cell(row = 22, column = 4).value = float(min_reserve.string)
    except ValueError:
        print('Mindanao NGCP data missing')
    
def date():
    if today.strftime('%A') == 'Monday':
        yesterday = today - datetime.timedelta(days = 3)
    else:
        yesterday = today - datetime.timedelta(days = 1)
    sheet.cell(row = 4, column =1).value = "Philippine Stock Exchange ({})".format(yesterday.strftime('%A, %d %B %Y'))
    sheet.cell(row = 18, column =1).value = "NGCP Power Situation Outlook ({})".format(today.strftime('%A, %d %B %Y'))


stockURL = {'http://www.bloomberg.com/quote/PHN:PM': 6,
            'http://www.bloomberg.com/quote/PHEN:PM': 7,
            'http://www.bloomberg.com/quote/PPG:PM':8,
            'http://www.bloomberg.com/quote/AP:PM': 9,
            'http://www.bloomberg.com/quote/ACR:PM': 10, 
            'http://www.bloomberg.com/quote/EDC:PM': 11, 
            'http://www.bloomberg.com/quote/FGEN:PM': 12, 
            'http://www.bloomberg.com/quote/MER:PM': 13, 
            'http://www.bloomberg.com/quote/SPC:PM': 14, 
            'http://www.bloomberg.com/quote/VVT:PM': 15
            }

for url in stockURL:
    stocks(url, stockURL[url])

forex()
index()
ngcp()
date()

wb.save(r'Daily_Energy_News_{}.xlsx'.format(today.strftime('%Y%m%d')))

ctypes.windll.user32.MessageBoxW(0, "It is finished", "Daily Energy News", 1)
